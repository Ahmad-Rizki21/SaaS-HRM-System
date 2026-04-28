import openpyxl

files = [
    r'c:\laragon\www\SaaS\PAYROLL\1. Feb.26-Judul perhituingan payrol.xlsx',
    r'c:\laragon\www\SaaS\PAYROLL\2. Contoh Slip gaji.xlsx',
    r'c:\laragon\www\SaaS\PAYROLL\3..Rekap gaji Gaji Jan.26.xlsx',
]

for fpath in files:
    print(f'\n{"="*80}')
    print(f'FILE: {fpath.split(chr(92))[-1]}')
    print(f'{"="*80}')
    wb = openpyxl.load_workbook(fpath, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'\n--- SHEET: {sheet_name} ---')
        print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
        
        # Also check merged cells
        if ws.merged_cells.ranges:
            print(f'Merged cells: {[str(m) for m in ws.merged_cells.ranges]}')
        
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), max_col=min(ws.max_column, 40), values_only=False):
            row_data = []
            for cell in row:
                if cell.value is not None:
                    row_data.append(f'{cell.coordinate}={repr(cell.value)}')
            if row_data:
                print(' | '.join(row_data))
